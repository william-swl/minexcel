import openpyxl as opx
import pandas as pd
from typing import Optional


def check_int_serial(ser, sort=False):
    # sort
    if sort:
        ser = sorted(list(ser))
    else:
        ser = list(ser)

    range_list = list(range(min(ser), max(ser) + 1))

    return ser == range_list


def read_excel_with_merged_cell(
    path: str, sheetname: Optional[str] = None
) -> pd.DataFrame:
    """Reads an Excel file and processes merged cells by filling the merged area with the top-left cell's value.

    Args:
        path: Path to the Excel file
        sheetname: Name of the sheet to read. If None, the active sheet is used

    Returns:
        DataFrame containing the data from the Excel sheet with merged cells processed

    Raises:
        AssertionError: If the provided sheetname is not in the workbook

    """
    wb = opx.load_workbook(path, data_only=True)
    if sheetname is None:
        ws = wb.active
    else:
        assert sheetname in wb.sheetnames
        ws = wb[sheetname]

    merged_ranges = [m for m in ws.merged_cells.ranges]

    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        cell_value = ws.cell(row=min_row, column=min_col).value

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = cell_value

    df = pd.DataFrame(ws.values, columns=None)

    return df
